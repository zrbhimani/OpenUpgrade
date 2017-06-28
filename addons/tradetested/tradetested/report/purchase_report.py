# -*- coding: utf-8 -*-
from odoo import tools, api, fields, models, _
import base64
import datetime

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


class purchase_report(models.Model):
    _inherit = 'purchase.report'

    month = fields.Selection([('01', 'January'), ('02', 'February'), ('03', 'March'), ('04', 'April'), ('05', 'May'), ('06', 'June'),
                              ('07', 'July'), ('08', 'August'), ('09', 'September'), ('10', 'October'), ('11', 'November'), ('12', 'December')], 'Month', readonly=True)
    location_id = fields.Many2one('stock.location', 'Destination', readonly=True)
    day = fields.Char('Day', size=128, readonly=True)
    name = fields.Char('Year', size=64, required=False, readonly=True)
    expected_date = fields.Date('Expected Date', readonly=True)
    validator = fields.Many2one('res.users', 'Validated By', readonly=True)


class report_po_shipments(models.TransientModel):
    _name = 'report.po.shipment'

    file = fields.Binary('File')
    file_name = fields.Char('File Name')

    @api.multi
    def get_report(self):
        sobj = self[0]
        today = datetime.date.today()
        start_date = today + datetime.timedelta(days=-today.weekday())
        end_date = start_date + datetime.timedelta(weeks=13)

        file_name = 'Purchase_order_incoming_shipments_' + start_date.strftime('%d_%b_%Y') + '.xlsx'

        self._cr.execute('''
            select
                pol.company_id,
                pt.categ_id,
                date_trunc('week', pol.date_planned)::date,
                sum(pol.product_qty * pol.price_unit) as sub_total
            from
                purchase_order_line pol
            LEFT JOIN
                product_product pp ON pol.product_id = pp.id
            LEFT JOIN
                product_template pt ON pt.id = pp.product_tmpl_id
            LEFT JOIN
                    product_category pc on pt.categ_id = pc.id
            LEFT JOIN
                    purchase_order po on pol.order_id = po.id
            WHERE
                po.state != 'cancel' AND
                pol.date_planned >= '%s' AND
                pol.date_planned <= '%s'
            group by 1,2,3
            order by 1,2,3
        ''' % (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))

        data = {}
        recs = self._cr.fetchall()
        for rec in recs:
            data[(rec[0], rec[1], rec[2])] = rec[3]

        headings = []
        while (start_date < end_date):
            headings.append(start_date)
            start_date = start_date + datetime.timedelta(weeks=1)

        categories = []
        categ_ids = self.env['product.category'].search([])
        for categ in categ_ids:
            categories.append([categ.id, categ.name])

        # Write in Excel
        wb = Workbook()
        sheet1 = wb.active
        sheet3 = wb.create_sheet()

        for company_id in [1, 3]:
            sheet = eval('sheet%s' % company_id)
            if company_id == 1:
                sheet.title = 'New Zealand'
                sheet.cell(row=1, column=1).value = 'New Zealand'
            elif company_id == 3:
                sheet.title = 'Australia'
                sheet.cell(row=1, column=1).value = 'Australia'

            sheet.cell(get_column_letter(1) + str(1)).font = Font(bold=True)
            sheet.column_dimensions[get_column_letter(1)].width = 30

            row_ind = 1
            col_ind = 2
            for heading in headings:
                sheet.cell(row=row_ind, column=col_ind).value = heading.strftime('%d-%b-%Y')
                col_ind += 1

            row_ind += 1
            for categ in categories + [[None, 'Un-Categorized']]:
                col_ind = 1
                sheet.cell(row=row_ind, column=col_ind).value = categ[1]

                for heading in headings:
                    col_ind += 1
                    sheet.cell(row=row_ind, column=col_ind).value = data.get((company_id, categ[0], heading.strftime('%Y-%m-%d')), '-')
                    sheet.cell(get_column_letter(col_ind) + str(row_ind)).number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
                row_ind += 1

            # Total
            row_ind + 1
            col_ind = 1
            sheet.cell(row=row_ind, column=col_ind).value = 'Total'
            for heading in headings:
                col_ind += 1
                sheet.cell(row=row_ind, column=col_ind).value = "=SUM(" + get_column_letter(col_ind) + str(2) + ':' + get_column_letter(col_ind) + str(row_ind - 1) + ")"
                sheet.cell(get_column_letter(col_ind) + str(row_ind)).number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'

        file_data = base64.encodestring(save_virtual_workbook(wb))

        sobj.write({'file': file_data, 'file_name': file_name})

        return {
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'report.po.shipment',
            'res_id': sobj.id,
            'type': 'ir.actions.act_window',
            'target': 'new'
        }